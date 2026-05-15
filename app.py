import streamlit as st
import pandas as pd
import numpy as np
import re
import glob
import shutil
from dotenv import load_dotenv
import OpenDartReader
import os
from datetime import datetime


# =========================
#  문자열 -> 숫자 변환 (강화)
# =========================
def to_number_strict(x):
    if pd.isna(x):
        return np.nan
    s = str(x)

    # 공백/제로폭/소프트하이픈 제거
    for ch in [
        "\u00a0",
        "\ufeff",
        "\u202f",
        "\u2009",
        "\u200a",
        "\u2007",
        "\u200b",
        "\u200c",
        "\u200d",
        "\u2060",
        "\u00ad",
    ]:
        s = s.replace(ch, "")

    # 통화/천단위 제거
    s = s.replace(",", "").replace("₩", "").replace("원", "").strip()

    # 하이픈/마이너스 통일
    s = (
        s.replace("\u2011", "-")  # non-breaking hyphen
        .replace("\u2212", "-")  # unicode minus
        .replace("–", "-")  # en dash
        .replace("—", "-")
    )  # em dash

    # 삼각형 음수표기 (△/▲) → 음수
    s = re.sub(r"^[\u25B3\u25B2]\s*", "-", s)

    # 괄호 음수표기: (1234) → -1234
    if re.fullmatch(r"\(.*\)", s):
        s = "-" + s[1:-1].strip()

    # 앞의 + 제거
    if s.startswith("+"):
        s = s[1:]

    # 숫자/부호/소수점 외 제거
    s = re.sub(r"[^0-9\-\.+]", "", s)

    if s in ("", "-", "--", "+"):
        return np.nan

    return pd.to_numeric(s, errors="coerce")


# =========================
#  엑셀 저장 (XlsxWriter + openpyxl 2차 교정)
# =========================
def save_excel_with_comma_format(df: pd.DataFrame, file_name: str):
    """
    1) 모든 '*amount' 열을 숫자형으로 보정
    2) XlsxWriter로 '셀 단위' 작성 (숫자→write_number, 결측→write_blank) + #,##0 서식
    3) openpyxl로 2차 검사/교정: 혹시 남은 문자열 셀은 숫자로 강제 변환 + #,##0 서식
    """
    import math, re
    from openpyxl import load_workbook

    def _norm_header(h: str) -> str:
        if h is None:
            return ""
        t = str(h).lower()
        for ch in [
            "\u00a0",
            "\ufeff",
            "\u202f",
            "\u2009",
            "\u200a",
            "\u2007",
            "\u200b",
            "\u200c",
            "\u200d",
            "\u2060",
            "\u00ad",
        ]:
            t = t.replace(ch, "")
        return t.strip()

    amount_cols = [c for c in df.columns if "amount" in _norm_header(c)]

    # 1) DF 숫자형 보정
    for col in amount_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # 2) XlsxWriter로 셀 단위 작성
    with pd.ExcelWriter(file_name, engine="xlsxwriter") as writer:
        wb = writer.book
        ws = wb.add_worksheet("Sheet1")
        writer.sheets["Sheet1"] = ws

        fmt_num = wb.add_format({"num_format": "#,##0"})
        fmt_text = wb.add_format()
        fmt_blank = wb.add_format({"num_format": "#,##0"})

        # 헤더
        for j, col in enumerate(df.columns):
            ws.write(0, j, col, fmt_text)

        # 데이터
        n_rows, n_cols = df.shape
        for i in range(n_rows):
            row = df.iloc[i]
            for j, col in enumerate(df.columns):
                val = row[col]
                if col in amount_cols:
                    if pd.isna(val):
                        ws.write_blank(i + 1, j, None, fmt_blank)
                    else:
                        ws.write_number(i + 1, j, float(val), fmt_num)
                else:
                    if pd.isna(val):
                        ws.write_blank(i + 1, j, None)
                    elif (
                        isinstance(val, (int, float))
                        and not isinstance(val, bool)
                        and math.isfinite(val)
                    ):
                        ws.write_number(i + 1, j, float(val))
                    else:
                        ws.write(i + 1, j, str(val))

        ws.autofilter(0, 0, n_rows, n_cols - 1)
        for j, col in enumerate(df.columns):
            if col in amount_cols:
                ws.set_column(j, j, 18, fmt_num)
            else:
                ws.set_column(j, j, 18)

    # 3) openpyxl로 2차 교정
    wb2 = load_workbook(file_name)
    ws2 = wb2.active

    # 헤더 인덱스 매핑
    header = [c.value for c in ws2[1]]
    name_to_colidx = {
        str(h): idx for idx, h in enumerate(header, start=1) if h is not None
    }

    def _to_number_strict_openpyxl(x):
        if x is None:
            return None
        s = str(x)
        for ch in [
            "\u00a0",
            "\ufeff",
            "\u202f",
            "\u2009",
            "\u200a",
            "\u2007",
            "\u200b",
            "\u200c",
            "\u200d",
            "\u2060",
            "\u00ad",
        ]:
            s = s.replace(ch, "")
        s = s.replace(",", "").replace("₩", "").replace("원", "").strip()
        s = (
            s.replace("\u2011", "-")
            .replace("\u2212", "-")
            .replace("–", "-")
            .replace("—", "-")
        )
        s = re.sub(r"^[\u25B3\u25B2]\s*", "-", s)
        if re.fullmatch(r"\(.*\)", s):
            s = "-" + s[1:-1].strip()
        if s.startswith("+"):
            s = s[1:]
        s = re.sub(r"[^0-9\-\.+]", "", s)
        if s in ("", "-", "--", "+"):
            return None
        try:
            return float(s)
        except Exception:
            return None

    for col in amount_cols:
        col_idx = name_to_colidx.get(col)
        if not col_idx:
            # normalize 이름으로 재탐색
            for h, idx in name_to_colidx.items():
                if "amount" in _norm_header(h) and _norm_header(h) == _norm_header(col):
                    col_idx = idx
                    break
        if not col_idx:
            continue

        for r in range(2, ws2.max_row + 1):
            cell = ws2.cell(row=r, column=col_idx)
            if isinstance(cell.value, str):
                num = _to_number_strict_openpyxl(cell.value)
                if num is not None:
                    cell.value = num
                    cell.number_format = "#,##0"
            else:
                if cell.value is not None:
                    cell.number_format = "#,##0"

    wb2.save(file_name)


# =========================
#  앱 본문
# =========================
load_dotenv()
api_key = os.getenv("DART_API_KEY") or st.secrets.get("DART_API_KEY", None)
if not api_key:
    api_key = st.sidebar.text_input("API 키를 입력하세요", type="password")
if not api_key:
    st.warning("API 키가 필요합니다. 입력 후 다시 시도하세요.")
    st.stop()

# 오늘 날짜 캐시 없으면 기존 캐시 재사용 (Streamlit Cloud 네트워크 타임아웃 방지)
_cache_dir = "docs_cache"
_today_cache = os.path.join(_cache_dir, f"opendartreader_corp_codes_{datetime.now().strftime('%Y%m%d')}.pkl")
if not os.path.exists(_today_cache):
    _existing = sorted(glob.glob(os.path.join(_cache_dir, "opendartreader_corp_codes_*.pkl")))
    if _existing:
        shutil.copy(_existing[-1], _today_cache)

@st.cache_resource
def init_dart(key):
    return OpenDartReader(key)

dart = init_dart(api_key)

st.title("📊 DART 재무제표 수집기")
st.markdown("종목코드를 입력하면 재무제표를 가져옵니다.")

# 종목코드 → 기업명 매핑
code_name_map = {
    "006400": "삼성SDI",
    "373220": "LG에너지솔루션",
    "01592447": "에스케이온",
    "259630": "엠플러스",
    "137400": "피엔티",
    "222080": "씨아이에스",
    "267320": "나인테크",
    "196490": "디에이테크놀로지",
    "109740": "디에스케이",
    "299030": "하나기술",
    "240600": "유진테크놀로지",
    "148930": "에이치와이티씨",
    "114810": "한솔아이원스",
    "137080": "나래나노텍"
}
company_names = list(code_name_map.values())

select_all = st.checkbox("✅ 전체 선택", value=True)
selected_names = st.multiselect(
    "조회할 기업 선택",
    options=company_names,
    default=company_names if select_all else [],
    key="corp_selector",
)
codes = [code for code, name in code_name_map.items() if name in selected_names]

current_year = datetime.now().year
year_range = list(range(current_year, current_year - 10, -1))
years = st.multiselect(
    "조회 연도 (복수 선택 가능)", year_range, default=[current_year - 1]
)

report_map = {
    "사업보고서": "11011",
    "반기보고서": "11012",
    "3분기보고서": "11014",
    "1분기보고서": "11013",
}
report_label = st.selectbox("보고서 유형", list(report_map.keys()))
report_code = report_map[report_label]

if st.button("📥 재무제표 수집"):
    if not codes:
        st.info("선택된 기업이 없습니다.")
        st.stop()
    if not years:
        st.info("선택된 연도가 없습니다.")
        st.stop()

    result_list = []
    for year in years:
        for code in codes:
            try:
                df = dart.finstate_all(code, bsns_year=year, reprt_code=report_code)
                if not (isinstance(df, pd.DataFrame) and not df.empty):
                    # XBRL 데이터 없을 경우 non-XBRL API로 fallback (연결→별도 순)
                    parts = []
                    for fs_div in ("CFS", "OFS"):
                        try:
                            fb = dart.finstate(code, year, report_code, fs_div=fs_div)
                            if isinstance(fb, pd.DataFrame) and not fb.empty:
                                parts.append(fb)
                        except Exception:
                            pass
                    df = pd.concat(parts, ignore_index=True) if parts else None

                if isinstance(df, pd.DataFrame) and not df.empty:
                    df["조회기업"] = code_name_map.get(code, code)
                    df["조회연도"] = year
                    result_list.append(df)
                    st.success(f"{year} - {code} 수집 완료")
                else:
                    st.warning(f"{year} - {code} 데이터 없음")
            except Exception as e:
                st.error(f"{year} - {code} 오류: {e}")

    if not result_list:
        st.info("수집된 데이터가 없습니다.")
        st.stop()

    result_df = pd.concat(result_list, ignore_index=True)

    # 모든 '*amount' 열 숫자화
    amount_like_cols = [c for c in result_df.columns if "amount" in str(c).lower()]
    for col in amount_like_cols:
        result_df[col] = result_df[col].apply(to_number_strict)
        result_df[col] = pd.to_numeric(result_df[col], errors="coerce").astype(
            "float64"
        )

    # 저장
    file_name = f"dart_finstate_{'_'.join(map(str, years))}.xlsx"
    save_excel_with_comma_format(result_df, file_name)

    with open(file_name, "rb") as f:
        st.download_button(
            label="📁 엑셀 다운로드",
            data=f,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.success("완료되었습니다! (모든 *amount 열 숫자형 + 엑셀 #,##0 포맷)")
