import os
import importlib

def test_package_imports():
    importlib.import_module("lib.expense")

def test_template_bundled():
    assert os.path.exists("templates/해외출장비정산서_v1.3.xlsx")

def test_template_has_expected_sheets():
    import openpyxl
    wb = openpyxl.load_workbook("templates/해외출장비정산서_v1.3.xlsx")
    assert "2-1출장경비 법인카드" in wb.sheetnames
    assert "출장비 정산(원화지급원화회수)" in wb.sheetnames
