import io
import pandas as pd
import openpyxl
from lib import expense

def test_end_to_end_haewoe():
    df = pd.DataFrame([
        {"승인일": "2026.06.27", "가맹점명": "ENMARKET 1330", "카드번호": "4074-6721-9739-3852",
         "이용자명": "공용", "현지이용금액": "41.65", "결제원금": "64,936"},
        {"승인일": "2026.06.02", "가맹점명": "FOOD LION #2811", "카드번호": "4074-6721-9739-3852",
         "이용자명": "공용", "현지이용금액": "541.78", "결제원금": "829,085"},
        {"승인일": "2026.06.05", "가맹점명": "WAWA 6309", "카드번호": "4074-6721-9739-3852",
         "이용자명": "공용", "현지이용금액": "44.00", "결제원금": "68,547"},
    ])
    rows = expense.normalize_haewoe(df)
    card = expense.extract_cards(rows)[0]
    rows = expense.filter_and_sort(rows, card)
    rows = expense.classify_rows(rows, expense.default_rules())
    # 마스터로 사용자명 주입
    for r in rows:
        r.user = "김동현"
    meta = expense.Meta(traveler="김동현", region="미국 조지아", usd_rate=1470, idr_rate=0.09)
    data = expense.build_settlement(rows, meta)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["2-1출장경비 법인카드"]
    # 날짜 오름차순: 06.02(FOOD LION) → 06.05(WAWA) → 06.27(ENMARKET)
    assert ws.cell(9, 3).value == "FOOD LION #2811"
    assert ws.cell(9, 4).value is None            # 마트 → 공란
    assert ws.cell(10, 3).value == "WAWA 6309"
    assert ws.cell(10, 4).value == "여비교통비"
    assert ws.cell(11, 4).value == "식비"          # ENMARKET
    assert ws.cell(12, 12).value == "=SUM(L9:L11)"
    # 원화 합계 검산
    total = sum(r.krw for r in rows)
    assert total == 64936 + 829085 + 68547
