import io
import openpyxl
from lib.expense.models import Row, Meta
from lib.expense.writer import build_settlement

def _rows(n):
    out = []
    for i in range(n):
        out.append(Row(date=f"2026.06.{i+1:02d}", shop=f"SHOP{i}", usd=10.0, idr=0,
                       krw=15000, user="김동현", source="haewoe", card_no="1",
                       category="식비", detail=""))
    return out

def test_build_fills_2_1_rows_and_total():
    rows = _rows(3)
    meta = Meta(traveler="김동현", start_date="2026.06.01", end_date="2026.06.30",
                region="미국 조지아", purpose="현장지원", usd_rate=1470, idr_rate=0.09)
    data = build_settlement(rows, meta)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["2-1출장경비 법인카드"]
    # 데이터 9~11, 합계 12
    assert ws.cell(9, 3).value == "SHOP0"          # C9 상호
    assert ws.cell(9, 12).value == 15000           # L9 원화금액
    assert ws.cell(9, 13).value == "김동현"         # M9 사용자
    assert ws.cell(9, 4).value == "식비"            # D9 항목
    assert ws.cell(9, 10).value == "=IF(H9>0,$L9/H9,0)"   # J9 수식
    assert ws.cell(12, 1).value == "합     계"
    assert ws.cell(12, 12).value == "=SUM(L9:L11)"

def test_build_writes_meta():
    data = build_settlement(_rows(1),
                            Meta(traveler="김동현", region="미국 조지아", usd_rate=1470, idr_rate=0.09))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["출장비 정산(원화지급원화회수)"]
    assert ws.cell(6, 4).value == "김동현"          # D6 출장자
    assert ws.cell(5, 9).value == 1470              # I5 USD 환율
    assert ws.cell(6, 9).value == 0.09              # I6 IDR 환율
