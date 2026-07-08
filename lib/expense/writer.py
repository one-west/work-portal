import io
import os
from copy import copy
import openpyxl

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "templates", "해외출장비정산서_v1.3.xlsx",
)

_SHEET_2_1 = "2-1출장경비 법인카드"
_SUMMARY_TABS = ["출장비 정산(원화지급원화회수)", "출장비 정산(외화지급외화회수)"]
_DATA_START = 9
_REF_DATA = 11     # 원본 데이터행(스타일 참조)
_REF_TOTAL = 31    # 원본 합계행(스타일 참조)

def _fill_2_1(ws, rows):
    n = len(rows)
    data_end = _DATA_START + n - 1
    total_row = data_end + 1

    rds = {c: copy(ws.cell(_REF_DATA, c)._style) for c in range(1, 14)}
    rdf = {c: ws.cell(_REF_DATA, c).number_format for c in range(1, 14)}
    rdh = ws.row_dimensions[_REF_DATA].height
    rts = {c: copy(ws.cell(_REF_TOTAL, c)._style) for c in range(1, 14)}
    rtf = {c: ws.cell(_REF_TOTAL, c).number_format for c in range(1, 14)}
    rth = ws.row_dimensions[_REF_TOTAL].height

    for m in list(ws.merged_cells.ranges):
        if m.min_row >= _DATA_START:
            ws.unmerge_cells(str(m))
    for r in range(_DATA_START, max(total_row + 2, _REF_TOTAL + 2)):
        for c in range(1, 15):
            ws.cell(r, c).value = None

    for i, row in enumerate(rows):
        r = _DATA_START + i
        ws.row_dimensions[r].height = rdh
        for c in range(1, 14):
            cell = ws.cell(r, c)
            cell._style = copy(rds[c])
            cell.number_format = rdf[c]
        ws.cell(r, 1).value = "=ROW()-8"
        ws.cell(r, 2).value = row.date
        ws.cell(r, 3).value = row.shop
        ws.cell(r, 4).value = row.category or None
        ws.cell(r, 5).value = "법인카드"
        ws.cell(r, 6).value = row.detail or None
        ws.cell(r, 8).value = float(row.usd)
        ws.cell(r, 9).value = float(row.idr)
        ws.cell(r, 10).value = f"=IF(H{r}>0,$L{r}/H{r},0)"
        ws.cell(r, 11).value = f"=IF(I{r}>0,$L{r}/I{r},0)"
        ws.cell(r, 12).value = int(round(row.krw))
        ws.cell(r, 13).value = row.user or "공용"
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

    ws.row_dimensions[total_row].height = rth
    for c in range(1, 14):
        cell = ws.cell(total_row, c)
        cell._style = copy(rts[c])
        cell.number_format = rtf[c]
    ws.cell(total_row, 1).value = "합     계"
    ws.cell(total_row, 8).value = f"=SUM(H{_DATA_START}:H{data_end})"
    ws.cell(total_row, 9).value = f"=SUM(I{_DATA_START}:I{data_end})"
    ws.cell(total_row, 12).value = f"=SUM(L{_DATA_START}:L{data_end})"
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    ws.merge_cells(start_row=total_row, start_column=6, end_row=total_row, end_column=7)

def _fill_meta(wb, meta):
    for name in _SUMMARY_TABS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        ws.cell(6, 4).value = meta.traveler        # D6 출장자
        ws.cell(7, 4).value = meta.start_date      # D7 기간 시작
        ws.cell(7, 5).value = meta.end_date        # E7 기간 종료
        ws.cell(8, 4).value = meta.region          # D8 지역/업체
        ws.cell(9, 4).value = meta.purpose         # D9 목적
        ws.cell(5, 9).value = meta.usd_rate        # I5 USD 환율
        ws.cell(6, 9).value = meta.idr_rate        # I6 IDR 환율

def build_settlement(rows, meta, template_path=TEMPLATE_PATH):
    wb = openpyxl.load_workbook(template_path)
    if _SHEET_2_1 not in wb.sheetnames:
        raise ValueError(f"템플릿에 '{_SHEET_2_1}' 시트가 없습니다.")
    _fill_2_1(wb[_SHEET_2_1], rows)
    _fill_meta(wb, meta)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
